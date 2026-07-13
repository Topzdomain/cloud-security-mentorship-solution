import re
import boto3
import json
import csv
import io
import argparse
from typing import List, Dict
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from rich.console import Console
from rich.table import Table
from rich import box
import PyPDF2  # type: ignore

console = Console()

# ─────────────────────────────────────────
#  PII Patterns
# ─────────────────────────────────────────
PII_PATTERNS = {
    'EMAIL':        (re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'), 'HIGH'),
    'US_SSN':       (re.compile(r'\b\d{3}-\d{2}-\d{4}\b'), 'CRITICAL'),
    'UK_NIN':       (re.compile(r'\b[A-Z]{2}\d{6}[A-D]\b'), 'CRITICAL'),
    'PHONE_US':     (re.compile(r'\b(\+1[-.\s]?)?\(?\d{3}\)?[-.\s]\d{3}[-.\s]\d{4}\b'), 'MEDIUM'),
    'CREDIT_CARD':  (re.compile(r'\b(?:4[0-9]{12}(?:[0-9]{3})?|5[1-5][0-9]{14}|3[47][0-9]{13})\b'), 'CRITICAL'),
    'AWS_KEY':      (re.compile(r'\bAKIA[0-9A-Z]{16}\b'), 'CRITICAL'),
    'PRIVATE_KEY':  (re.compile(r'-----BEGIN (RSA |EC )?PRIVATE KEY-----'), 'CRITICAL'),
    'IP_ADDRESS':   (re.compile(r'\b(?:10|172\.(?:1[6-9]|2[0-9]|3[01])|192\.168)\.\d{1,3}\.\d{1,3}\b'), 'LOW'),
    'FULL_NAME':    (re.compile(r'\b[A-Z][a-z]{1,20}\s[A-Z][a-z]{1,20}(?:\s[A-Z][a-z]{1,20})?\b'), 'MEDIUM'),
    'DATE_OF_BIRTH':(re.compile(
                        r'\b(?:'
                        r'\d{2}[\/\-\.]\d{2}[\/\-\.]\d{4}'
                        r'|\d{4}[\/\-\.]\d{2}[\/\-\.]\d{2}'
                        r'|(?:0?[1-9]|[12][0-9]|3[01])\s'
                        r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*'
                        r'\s\d{4}'
                        r')\b'), 'HIGH'),
}

BINARY_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.zip', '.gz', '.tar', '.mp4', '.exe', '.bin'}
SEVERITY_ORDER    = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}


# ─────────────────────────────────────────
#  Deduplication Helper
# ─────────────────────────────────────────
def _deduplicate(raw_findings: List[Dict]) -> List[Dict]:
    """
    Collapse multiple hits of the same PII type in the same file into one row.
    Aggregates: total match count, all unique locations, one masked sample.
    Result: one row per (bucket, object_key, pii_type).
    """
    grouped = defaultdict(lambda: {
        'match_count': 0,
        'locations':   [],
        'sample':      None,
        'severity':    None,
        'bucket':      None,
        'object_key':  None,
        'pii_type':    None,
    })

    for f in raw_findings:
        key = (f['bucket'], f['object_key'], f['pii_type'])
        g = grouped[key]
        g['match_count'] += f['match_count']
        # Only capture the first location — where the sample was found
        if g['sample'] is None:
            g['sample']     = f['sample']
            g['severity']   = f['severity']
            g['bucket']     = f['bucket']
            g['object_key'] = f['object_key']
            g['pii_type']   = f['pii_type']
            g['location']   = f.get('location', 'N/A')

    results = []
    for g in grouped.values():
        results.append({
            'bucket':      g['bucket'],
            'object_key':  g['object_key'],
            'pii_type':    g['pii_type'],
            'severity':    g['severity'],
            'match_count': g['match_count'],
            'location':    g.get('location', 'N/A'),
            'sample':      g['sample'],
        })

    return sorted(results, key=lambda x: SEVERITY_ORDER.get(x['severity'], 9))


# ─────────────────────────────────────────
#  Location-Aware PII Scanners (per file type)
# ─────────────────────────────────────────
def _scan_text(content: str) -> List[Dict]:
    """Scan plain text line by line — returns line number as location."""
    findings = []
    for line_no, line in enumerate(content.splitlines(), start=1):
        for pii_type, (pattern, severity) in PII_PATTERNS.items():
            matches = pattern.findall(line)
            if matches:
                findings.append({
                    'pii_type':    pii_type,
                    'severity':    severity,
                    'match_count': len(matches),
                    'sample':      _mask(str(matches[0])),
                    'location':    f"Line {line_no}",
                })
    return findings


def _scan_csv(content: str) -> List[Dict]:
    """Scan CSV row by row — returns Row number + Column header as location."""
    findings = []
    try:
        reader = csv.DictReader(io.StringIO(content))
        for row_no, row in enumerate(reader, start=2):
            for col_name, cell_value in row.items():
                if not cell_value:
                    continue
                for pii_type, (pattern, severity) in PII_PATTERNS.items():
                    matches = pattern.findall(str(cell_value))
                    if matches:
                        findings.append({
                            'pii_type':    pii_type,
                            'severity':    severity,
                            'match_count': len(matches),
                            'sample':      _mask(str(matches[0])),
                            'location':    f"Row {row_no}, Col '{col_name}'",
                        })
    except Exception:
        findings = _scan_text(content)
    return findings


def _scan_json(content: str) -> List[Dict]:
    """Scan JSON values — returns dot-notation key path as location."""
    findings = []
    try:
        data = json.loads(content)
        _scan_json_node(data, path='', findings=findings)
    except Exception:
        findings = _scan_text(content)
    return findings


def _scan_json_node(node, path: str, findings: list):
    if isinstance(node, dict):
        for key, value in node.items():
            _scan_json_node(value, path=f"{path}.{key}" if path else key, findings=findings)
    elif isinstance(node, list):
        for idx, item in enumerate(node):
            _scan_json_node(item, path=f"{path}[{idx}]", findings=findings)
    elif isinstance(node, str):
        for pii_type, (pattern, severity) in PII_PATTERNS.items():
            matches = pattern.findall(node)
            if matches:
                findings.append({
                    'pii_type':    pii_type,
                    'severity':    severity,
                    'match_count': len(matches),
                    'sample':      _mask(str(matches[0])),
                    'location':    f"Key: {path}",
                })


def _scan_pdf(pdf_bytes: bytes) -> List[Dict]:
    """Scan PDF page by page — returns page number as location."""
    findings = []
    try:
        reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
        for page_no, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ''
            for pii_type, (pattern, severity) in PII_PATTERNS.items():
                matches = pattern.findall(text)
                if matches:
                    findings.append({
                        'pii_type':    pii_type,
                        'severity':    severity,
                        'match_count': len(matches),
                        'sample':      _mask(str(matches[0])),
                        'location':    f"Page {page_no}",
                    })
    except Exception as e:
        console.print(f"    [yellow][!] PDF extraction failed: {e}[/yellow]")
    return findings


def _scan_excel(excel_bytes: bytes) -> List[Dict]:
    """Scan Excel file cell by cell — returns Sheet + Cell reference as location."""
    findings = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    cell_str = str(cell.value)
                    for pii_type, (pattern, severity) in PII_PATTERNS.items():
                        matches = pattern.findall(cell_str)
                        if matches:
                            findings.append({
                                'pii_type':    pii_type,
                                'severity':    severity,
                                'match_count': len(matches),
                                'sample':      _mask(str(matches[0])),
                                'location':    f"Sheet '{sheet_name}', Cell {cell.coordinate}",
                            })
        wb.close()
    except ImportError:
        console.print("    [yellow][!] openpyxl not installed — skipping Excel file. Run: pip install openpyxl[/yellow]")
    except Exception as e:
        console.print(f"    [yellow][!] Excel scan failed: {e}[/yellow]")
    return findings


# ─────────────────────────────────────────
#  Core Scanner
# ─────────────────────────────────────────
def scan_bucket_for_pii(bucket_name: str, max_objects: int = 50,
                         max_bytes_per_object: int = 512000) -> List[Dict]:
    """Sample objects from a bucket and scan for PII patterns."""
    s3 = boto3.client('s3')
    raw_findings = []

    try:
        paginator = s3.get_paginator('list_objects_v2')
        pages = paginator.paginate(Bucket=bucket_name, PaginationConfig={'MaxItems': max_objects})

        for page in pages:
            for obj in page.get('Contents', []):
                key       = obj['Key']
                size_kb   = round(obj.get('Size', 0) / 1024, 1)
                key_lower = key.lower()

                try:
                    if key_lower.endswith('.pdf'):
                        console.print(f"    Scanning PDF:   [dim]{key}[/dim] ({size_kb} KB)")
                        raw     = s3.get_object(Bucket=bucket_name, Key=key)['Body'].read(max_bytes_per_object)
                        results = _scan_pdf(raw)

                    elif key_lower.endswith('.csv'):
                        console.print(f"    Scanning CSV:   [dim]{key}[/dim] ({size_kb} KB)")
                        raw     = s3.get_object(Bucket=bucket_name, Key=key)['Body'].read(max_bytes_per_object)
                        results = _scan_csv(raw.decode('utf-8', errors='ignore'))

                    elif key_lower.endswith('.json'):
                        console.print(f"    Scanning JSON:  [dim]{key}[/dim] ({size_kb} KB)")
                        raw     = s3.get_object(Bucket=bucket_name, Key=key)['Body'].read(max_bytes_per_object)
                        results = _scan_json(raw.decode('utf-8', errors='ignore'))

                    elif key_lower.endswith(('.xlsx', '.xls')):
                        console.print(f"    Scanning Excel: [dim]{key}[/dim] ({size_kb} KB)")
                        raw     = s3.get_object(Bucket=bucket_name, Key=key)['Body'].read(max_bytes_per_object)
                        results = _scan_excel(raw)

                    elif any(key_lower.endswith(ext) for ext in BINARY_EXTENSIONS):
                        console.print(f"    [dim]Skipping binary: {key}[/dim]")
                        continue

                    else:
                        console.print(f"    Scanning text:  [dim]{key}[/dim] ({size_kb} KB)")
                        raw     = s3.get_object(Bucket=bucket_name, Key=key)['Body'].read(max_bytes_per_object)
                        results = _scan_text(raw.decode('utf-8', errors='ignore'))

                    # Attach bucket + object key then collect
                    for r in results:
                        r['bucket']     = bucket_name
                        r['object_key'] = key
                    raw_findings.extend(results)

                except Exception as e:
                    console.print(f"    [yellow][!] Could not read {key}: {e}[/yellow]")
                    continue

    except Exception as e:
        console.print(f"  [red][!] Could not scan {bucket_name}: {e}[/red]")

    # Deduplicate: one row per (bucket, object, pii_type)
    return _deduplicate(raw_findings)


def scan_all_buckets_for_pii(profile: str = None, max_objects: int = 50,
                               max_bytes: int = 512000) -> Dict[str, List[Dict]]:
    """Scan all accessible S3 buckets for PII."""
    session  = boto3.Session(profile_name=profile) if profile else boto3.Session()
    s3       = session.client('s3')
    buckets  = s3.list_buckets().get('Buckets', [])
    all_results = {}

    for bucket in buckets:
        name = bucket['Name']
        console.print(f"\n  [cyan]Scanning bucket:[/cyan] {name}")
        all_results[name] = scan_bucket_for_pii(name, max_objects, max_bytes)

    return all_results


# ─────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────
def _mask(value: str) -> str:
    if len(value) <= 4:
        return '****'
    return value[:2] + '*' * (len(value) - 4) + value[-2:]


# ─────────────────────────────────────────
#  Output — Console Table
# ─────────────────────────────────────────
def print_findings_table(all_results: Dict[str, List[Dict]]):
    all_findings = [f for findings in all_results.values() for f in findings]

    if not all_findings:
        console.print("\n[green]✅ No PII detected across scanned buckets.[/green]")
        return

    table = Table(title="PII Detection Findings", box=box.ROUNDED, show_lines=True)
    table.add_column("Severity",  style="bold", width=10)
    table.add_column("Bucket",    width=25)
    table.add_column("Object",    width=20)
    table.add_column("PII Type",  width=15)
    table.add_column("Locations", width=30)
    table.add_column("Count",     width=6)
    table.add_column("Sample",    width=18)

    severity_styles = {
        'CRITICAL': 'red', 'HIGH': 'orange3', 'MEDIUM': 'yellow', 'LOW': 'green'
    }

    for f in all_findings:
        style = severity_styles.get(f['severity'], 'white')
        table.add_row(
            f"[{style}]{f['severity']}[/{style}]",
            f['bucket'],
            f['object_key'],
            f['pii_type'],
            f.get('location', 'N/A'),
            str(f['match_count']),
            f['sample']
        )

    console.print(table)


# ─────────────────────────────────────────
#  Output — JSON Report
# ─────────────────────────────────────────
def generate_report(all_results: Dict[str, List[Dict]], output_dir: str = 'reports') -> dict:
    Path(output_dir).mkdir(exist_ok=True)
    timestamp    = datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S')
    all_findings = [f for findings in all_results.values() for f in findings]

    report = {
        'report_metadata': {
            'timestamp':              datetime.utcnow().isoformat(),
            'tool':                   'S3 PII Detector',
            'total_buckets_scanned':  len(all_results),
            'total_objects_with_pii': len({f['object_key'] for f in all_findings}),
        },
        'summary': {
            'total_findings': len(all_findings),
            'critical':       sum(1 for f in all_findings if f['severity'] == 'CRITICAL'),
            'high':           sum(1 for f in all_findings if f['severity'] == 'HIGH'),
            'medium':         sum(1 for f in all_findings if f['severity'] == 'MEDIUM'),
            'low':            sum(1 for f in all_findings if f['severity'] == 'LOW'),
        },
        'findings_by_bucket': {
            bucket: findings
            for bucket, findings in all_results.items()
            if findings
        }
    }

    json_path = f"{output_dir}/{timestamp}-pii-scan.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2)

    console.print(f"\n[+] JSON report saved: [green]{json_path}[/green]")
    return report


# ─────────────────────────────────────────
#  Entry Point
# ─────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='S3 PII Detector')
    parser.add_argument('--bucket',      default=None,   help='Scan a specific bucket only')
    parser.add_argument('--profile',     default=None,   help='AWS profile to use')
    parser.add_argument('--max-objects', default=50,     type=int, help='Max objects per bucket (default: 50)')
    parser.add_argument('--max-bytes',   default=512000, type=int, help='Max bytes per object (default: 500KB)')
    parser.add_argument('--output',      default='reports', help='Output directory for reports')
    args = parser.parse_args()

    console.print("\n[bold blue]🔍 S3 PII Detector[/bold blue]\n")
    console.print(f"Max objects per bucket: [yellow]{args.max_objects}[/yellow]")
    console.print(f"Max bytes per object:   [yellow]{args.max_bytes}[/yellow]\n")

    if args.bucket:
        console.print(f"[*] Scanning bucket: [cyan]{args.bucket}[/cyan]")
        findings    = scan_bucket_for_pii(args.bucket, args.max_objects, args.max_bytes)
        all_results = {args.bucket: findings}
    else:
        console.print("[*] Scanning all accessible S3 buckets...", style="cyan")
        all_results = scan_all_buckets_for_pii(args.profile, args.max_objects, args.max_bytes)

    print_findings_table(all_results)
    report = generate_report(all_results, args.output)

    console.print(f"\n[bold green]✅ PII Scan Complete![/bold green]")
    console.print(f"   Buckets scanned:        [white]{report['report_metadata']['total_buckets_scanned']}[/white]")
    console.print(f"   Objects with PII:       [white]{report['report_metadata']['total_objects_with_pii']}[/white]")
    console.print(f"   Critical:               [red]{report['summary']['critical']}[/red]")
    console.print(f"   High:                   [orange3]{report['summary']['high']}[/orange3]")
    console.print(f"   Medium:                 [yellow]{report['summary']['medium']}[/yellow]")
    console.print(f"   Low:                    [green]{report['summary']['low']}[/green]")


if __name__ == '__main__':
    main()